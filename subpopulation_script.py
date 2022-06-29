import sqlite3
import os
import sys
import openpyxl as px
import numpy as np
import pandas as pd
from datetime import datetime

# CONSTANTS
TOP = "K8"  
LM = "K9"  
MM = "L9"   
RM = "M9" 
LO = "K10"
MO = "L10"
RO = "M10"
LLO = "K11"
LRO = "L11"
MLO = "M11"
MRO = "N11"
RLO = "O11"
RRO = "P11"

NB_ITERS = "C2"
NB_PARENTS = "C3"
NB_BEST = "C4"
CR = "C5"
MR = "C6"


jar_file = "\"/Users/dennismaes/Library/Mobile Documents/com~apple~CloudDocs/School/2021-2022/Thesis/applicatie/thesis/target/scheduler-1.0.0.jar:/Users/dennismaes/.m2/repository/org/xerial/sqlite-jdbc/3.36.0/sqlite-jdbc-3.36.0.jar:/Users/dennismaes/.m2/repository/org/javatuples/javatuples/1.2/javatuples-1.2.jar\""
rpath = "../scheduler/backend/generator_test3.db"

conn = sqlite3.connect(rpath)
cur = conn.cursor()


RESULTS_FILE= "one_point_test6.xlsx"


wknd_shifts = ["JAWE", "SAWE", "TPWE"]
week_shifts = ["SANW", "JAEV"]
holiday_shifts = ["JAHO", "SAHO", "TPHO"]
friday_shifts = ["TPNF", "SAEV2"]
mo_do_shifts = ["SAEV1"]

"""WL = dict()
for sh in wknd_shifts:
    WL[sh] = 1.5
for sh in week_shifts:
    WL[sh] = 1

WL["JAHO"] = 1.5
WL["SAHO"] = 2
WL["TPHO"] = 3"""

    
def parse_excel_input(ws):
    return ws[TOP].value, ws[LM].value, ws[MM].value, ws[RM].value, ws[LO].value, ws[MO].value, ws[RO].value, ws[LLO].value, ws[LRO].value, ws[MLO].value, ws[MRO].value, ws[RLO].value, ws[RRO].value

def parse_excel_input_smaller(ws):
    return ws[TOP].value, ws[LM].value, ws[MM].value, ws[RM].value, ws[LO].value, ws[MO].value, ws[RO].value

def parse_excel_input_no_subpopulations(ws):
    return [ws[TOP].value]

def parse_params(wb):
    ws = wb["params"]
    return ws[NB_ITERS].value, ws[NB_PARENTS].value, ws[NB_BEST].value, ws[CR].value, ws[MR].value


def write_to_excel(results_dict, filename=None):
    if filename is None:
        wb = px.load_workbook(RESULTS_FILE)
    else:
        wb = px.load_workbook(filename)
    print(wb.worksheets)
    wb.create_sheet(index=len(wb.sheetnames), title="results "+  str(len(wb.worksheets)))
    ws = wb["results "+  str(len(wb.worksheets) -1)]
    ws["C2"] = "Times < 12 rest days"
    ws["D2"] = "Total WL"
    row = 3
    for x in results_dict.keys():
        ws["B" + str(row)] = x
        ws["C" + str(row)] = results_dict[x][0]
        ws["D" + str(row)] = results_dict[x][1]
        row += 1
    
    ws["B" + str(row + 1)] = "TOTAL"
    ws["C" + str(row + 1)] = get_sum(results)
    ws["D" + str(row + 1)] = get_fair(results)
    
    ws["B" + str(row + 2)] = "STD DEV"
    ws["D" + str(row + 2)] = get_std(results)

    if filename is None:
        wb.save(RESULTS_FILE)
    else:
        wb.save(filename)

    
def write_to_excel_ws(results_dict, wb, ws, filename=None):
    ws["C2"] = "Times <= 7 rest days"
    ws["D2"] = "Times < 12 rest days"
    ws["E2"] = "Times <= 14 rest days"
    ws["F2"] = "Times <= 21 rest days"
    ws["G2"] = "Total WL"
    row = 3
    for x in results_dict[0].keys():
        ws["B" + str(row)] = x
        ws["C" + str(row)] = results_dict[0][x][0]
        ws["D" + str(row)] = results_dict[1][x][0]
        ws["E" + str(row)] = results_dict[2][x][0]
        ws["F" + str(row)] = results_dict[3][x][0]
        ws["G" + str(row)] = results_dict[0][x][1]
        row += 1
    
    ws["B" + str(row + 1)] = "TOTAL"
    ws["C" + str(row + 1)] = get_sum(results_dict[0])
    ws["D" + str(row + 1)] = get_sum(results_dict[1])
    ws["E" + str(row + 1)] = get_sum(results_dict[2])
    ws["F" + str(row + 1)] = get_sum(results_dict[3])
    ws["G" + str(row + 1)] = get_fair(results_dict[0])
    
    ws["B" + str(row + 2)] = "STD DEV"
    ws["G" + str(row + 2)] = get_std(results_dict[0])

    if filename is None:
        wb.save(RESULTS_FILE)
    else:
        wb.save(filename)
    
def get_nb_times_rest_less_than(assignments, min_bal=12):
    first_shift = True
    counter = 0
    total = 0
    for ass in assignments:
        if ass == "FREE":
            if not first_shift:
                counter += 1
        else:
            if not first_shift:
                if counter < min_bal and counter > 0:
                    total += 1
            else: 
                first_shift = False
            
            counter = 0
    return total


def get_workloads():
    result = dict()
    for row in cur.execute('''SELECT shift_type, shift_workload FROM shift_type_parameters'''):
        result[row[0]] = row[1]
    
    return result


def get_total_wl(assignments):
    wl = 0.0
    WL = get_workloads()
    #print([x for x in str(assignments).replace("[","").replace("]", "").replace(",","").replace('FREE', "|").split("|") if x != "\'" and x != "\' \'"])
    for sh in [x for x in str(assignments).replace("[","").replace("]", "").replace(",","").replace('FREE', "|").split("|") if x != "\'" and x != "\' \'"]:
    #    print(sh)
        # wl += WL[sh.split(" ")[1].replace("\'", "")]
        if sh.split(" ")[0].replace("\'", "") in WL.keys():
            wl += WL[sh.split(" ")[0].replace("\'", "")]
        #if len(sh.split(" ")) == 1:
         #   wl += WL[sh.split(" ")[0].replace("\'", "")]
        else:
            wl += WL[sh.split(" ")[1].replace("\'", "")]
    return wl

def get_std(results):
    return np.std([results[x][1] for x in results.keys()])


def get_sum(results):
    return np.sum([results[x][0] for x in results.keys() ])

def get_fair(results):
    return max([results[x][1] for x in results.keys()]) - min([results[x][1] for x in results.keys()])


def read_results(filename):
    with open(filename, 'r') as f:
        line = f.readlines()[-1]
        splitted = line.split("|")
        best = float(splitted[0].split(":")[1])
        coverage = float(splitted[1].split(":")[1])
        balance = float(splitted[2].split(":")[1])
        fairness = float(splitted[3].replace("\n", "").split(":")[1])
        avg = float(splitted[4].replace("\n", "").split(":")[1])
        return best, coverage, balance, fairness, avg

def read_results2(filename):

    df = pd.DataFrame()

    l_best = list()
    l_coverage = list()
    l_balance = list()
    l_fairness = list()
    l_avg = list()
    l_times = list()
    with open(filename, 'r') as f:

        for line in f.readlines():
            if line[0] == "B":
                splitted = line.split("|")
                best = float(splitted[0].split(":")[1])
                coverage = float(splitted[1].split(":")[1])
                balance = float(splitted[2].split(":")[1])
                fairness = float(splitted[3].replace("\n", "").split(":")[1])
                avg = float(splitted[4].replace("\n", "").split(":")[1])
                time = datetime.strptime(splitted[5].replace("\n", "").split(":")[1], " %H/%M/%S")

                l_best.append(best)
                l_coverage.append(coverage)
                l_balance.append(balance)
                l_fairness.append(fairness)
                l_avg.append(avg)
                l_times.append(time)

        df["best"] = l_best
        df["coverage"] = l_coverage
        df["balance"] = l_balance
        df["fairness"] = l_fairness
        df["avg"] = l_avg
        df["time"] = l_times


        return l_best[-1], l_coverage[-1], l_balance[-1], l_fairness[-1], l_avg[-1], df


def run_script():
    wb = px.load_workbook(RESULTS_FILE)
    for sh_n in wb.sheetnames:
        print("Current sheetname: ", sh_n)
        ws = wb[sh_n]
        input_data = parse_excel_input(ws)
        
        run_string = "java -cp " + jar_file + " Main"

        for x in input_data:
            run_string += " \"{}\"".format(x)
        

        run_string += " > tmp.txt"

        os.system(run_string)
        
        schedule_entries = []
        for row in cur.execute('''SELECT assistant_id, day_nb, shift_type FROM assignment'''):
            schedule_entries.append(row)

        assistant_entries = dict()
        for row in cur.execute('''SELECT id, name FROM assistant'''):
            assistant_entries[row[0]] = row[1].strip()
        
        results = dict()
        for as_id in list(assistant_entries.keys()):
            current_assignments = [x for x in schedule_entries if x[0] == as_id]
            current_name = assistant_entries[as_id]
            shifts = [ x[2] for x in current_assignments]
            i = 0
            free_cnt = 0
            while i < len(shifts):
                if shifts[i] == 'FREE':
                    free_cnt += 1
                else:
                    if free_cnt == 2:
                        shifts[i-2] = 'JAEV'
                        shifts[i-1] = 'JAEV'
                    free_cnt = 0
                i += 1

            results[current_name] = (get_nb_times_rest_less_than(shifts), get_total_wl(shifts))
        write_to_excel_ws(results, wb, ws)
        stat = read_results("tmp.txt")

        ws["F2"] = "FITNESS"
        ws["G2"] = "COVERAGE"
        ws["H2"] = "BALANCE"
        ws["I2"] = "FAIRNESS"
        ws["J2"] = "AVG"

        ws["F3"] = stat[0]
        ws["G3"] = stat[1]
        ws["H3"] = stat[2]
        ws["I3"] = stat[3]
        ws["J3"] = stat[4]

        wb.save(RESULTS_FILE)


def run_script2():
    wb = px.load_workbook(RESULTS_FILE)
    sh_n = "Blad1"
    print("Current sheetname: ", sh_n)
    ws = wb[sh_n]
    input_data = parse_excel_input(ws)

    run_string = "java -cp " + jar_file + " Main"

    for x in input_data:
        run_string += " \"{}\"".format(x)

    input_params = parse_params(wb)
    for x in input_params:
        run_string += " \"{}\"".format(x)

    print(run_string)
    run_string += " > tmp.txt"

    os.system(run_string)

    schedule_entries = []
    for row in cur.execute('''SELECT assistant_id, day_nb, shift_type FROM assignment'''):
        schedule_entries.append(row)

    assistant_entries = dict()
    for row in cur.execute('''SELECT id, name FROM assistant'''):
        assistant_entries[row[0]] = row[1].strip()
    
    results = dict()
    for as_id in list(assistant_entries.keys()):
        current_assignments = [x for x in schedule_entries if x[0] == as_id]
        current_name = assistant_entries[as_id]
        shifts = [x[2] for x in current_assignments]
        i = 0
        free_cnt = 0
        while i < len(shifts):
            if shifts[i] == 'FREE':
                free_cnt += 1
            else:
                if free_cnt == 2:
                    shifts[i - 2] = 'JAEV'
                    shifts[i - 1] = 'JAEV'
                free_cnt = 0
            i += 1

        results[current_name] = (get_nb_times_rest_less_than(shifts), get_total_wl(shifts))
    write_to_excel_ws(results, wb, ws)
    stat = read_results2("tmp.txt")

    ws["H2"] = "FITNESS"
    ws["I2"] = "COVERAGE"
    ws["J2"] = "BALANCE"
    ws["K2"] = "FAIRNESS"
    ws["L2"] = "AVG"

    ws["H3"] = stat[0]
    ws["I3"] = stat[1]
    ws["J3"] = stat[2]
    ws["K3"] = stat[3]
    ws["L3"] = stat[4]


    wb.save(RESULTS_FILE)

    with pd.ExcelWriter(RESULTS_FILE, mode='a') as writer:
        stat[5].to_excel(writer, sheet_name='Blad2')


def run_all(dir_name, smaller=True):
    arr = os.listdir(dir_name)
    arr = [a for a in arr if a[0:3] == 'one']
    counter = 0
    for res_file in arr:
        counter += 1
        res_file = dir_name + res_file
        print("CURRENT FILE: ", res_file)
        print("STATUS: ", counter, "/", len(arr))

        wb = px.load_workbook(res_file)
        sh_n = "Blad1"
        print("Current sheetname: ", sh_n)
        ws = wb[sh_n]

        if smaller:
            input_data = parse_excel_input_smaller(ws)
        else:
            input_data = parse_excel_input(ws)

        run_string = "java -cp " + jar_file + " Main"

        for x in input_data:
            run_string += " \"{}\"".format(x)

        input_params = parse_params(wb)
        for x in input_params:
            run_string += " \"{}\"".format(x)

        print(run_string)
        run_string += " > tmp.txt"

        os.system(run_string)

        

        schedule_entries = []
        for row in cur.execute('''SELECT assistant_id, day_nb, shift_type FROM assignment'''):
            schedule_entries.append(row)

        assistant_entries = dict()
        for row in cur.execute('''SELECT id, name FROM assistant'''):
            assistant_entries[row[0]] = row[1].strip()

        results_list = list()
        results7 = dict()
        results12 = dict()
        results14 = dict()
        results21 = dict()
        for as_id in list(assistant_entries.keys()):
            current_assignments = [x for x in schedule_entries if x[0] == as_id]
            current_name = assistant_entries[as_id]
            shifts = [x[2] for x in current_assignments]
            i = 0
            free_cnt = 0
            while i < len(shifts):
                if shifts[i] == 'FREE':
                    free_cnt += 1
                else:
                    if free_cnt == 2:
                        shifts[i - 2] = 'JAEV'
                        shifts[i - 1] = 'JAEV'
                    free_cnt = 0
                i += 1

            results7[current_name] = (get_nb_times_rest_less_than(shifts, 7), get_total_wl(shifts))
            results12[current_name] = (get_nb_times_rest_less_than(shifts, 12), get_total_wl(shifts))
            results14[current_name] = (get_nb_times_rest_less_than(shifts, 14), get_total_wl(shifts))
            results21[current_name] = (get_nb_times_rest_less_than(shifts, 21), get_total_wl(shifts))
        results_list.append(results7)
        results_list.append(results12)
        results_list.append(results14)
        results_list.append(results21)

        write_to_excel_ws(results_list, wb, ws, res_file)
        stat = read_results2("tmp.txt")

        ws["I2"] = "FITNESS"
        ws["J2"] = "COVERAGE"
        ws["K2"] = "BALANCE"
        ws["L2"] = "FAIRNESS"
        ws["M2"] = "AVG"

        ws["I3"] = stat[0]
        ws["J3"] = stat[1]
        ws["K3"] = stat[2]
        ws["L3"] = stat[3]
        ws["M3"] = stat[4]

        wb.save(res_file)

        with pd.ExcelWriter(res_file, mode='a') as writer:
            stat[5].to_excel(writer, sheet_name='Blad2')


def run_all_custom_db(dir_name, db_name, smaller=True):
    
    arr = os.listdir(dir_name)
    arr = [a for a in arr if a[0:3] == 'one']
    counter = 0
    for res_file in arr:
        counter += 1
        res_file = dir_name + res_file
        print("CURRENT FILE: ", res_file)
        print("STATUS: ", counter, "/", len(arr))

        wb = px.load_workbook(res_file)
        sh_n = "Blad1"
        print("Current sheetname: ", sh_n)
        ws = wb[sh_n]

        if smaller:
            input_data = parse_excel_input_smaller(ws)
        else:
            input_data = parse_excel_input(ws)

        run_string = "java -cp " + jar_file + " Main"

        for x in input_data:
            run_string += " \"{}\"".format(x)

        input_params = parse_params(wb)
        for x in input_params:
            run_string += " \"{}\"".format(x)

        run_string += " "
        run_string += " \"" + db_name + "\"" 

        print(run_string)
        run_string += " > tmp.txt"

        os.system(run_string)
        db_name_full = "/Users/dennismaes" + db_name
       
        conn = sqlite3.connect(db_name_full)
        cur = conn.cursor()

        schedule_entries = []
        for row in cur.execute('''SELECT assistant_id, day_nb, shift_type FROM assignment'''):
            schedule_entries.append(row)

        assistant_entries = dict()
        for row in cur.execute('''SELECT id, name FROM assistant'''):
            assistant_entries[row[0]] = row[1].strip()

        results_list = list()
        results7 = dict()
        results12 = dict()
        results14 = dict()
        results21 = dict()
        for as_id in list(assistant_entries.keys()):
            current_assignments = [x for x in schedule_entries if x[0] == as_id]
            current_name = assistant_entries[as_id]
            shifts = [x[2] for x in current_assignments]
            i = 0
            free_cnt = 0
            while i < len(shifts):
                if shifts[i] == 'FREE':
                    free_cnt += 1
                else:
                    if free_cnt == 2:
                        shifts[i - 2] = 'JAEV'
                        shifts[i - 1] = 'JAEV'
                    free_cnt = 0
                i += 1

            results7[current_name] = (get_nb_times_rest_less_than(shifts, 7), get_total_wl(shifts))
            results12[current_name] = (get_nb_times_rest_less_than(shifts, 12), get_total_wl(shifts))
            results14[current_name] = (get_nb_times_rest_less_than(shifts, 14), get_total_wl(shifts))
            results21[current_name] = (get_nb_times_rest_less_than(shifts, 21), get_total_wl(shifts))
        results_list.append(results7)
        results_list.append(results12)
        results_list.append(results14)
        results_list.append(results21)

        write_to_excel_ws(results_list, wb, ws, res_file)
        stat = read_results2("tmp.txt")

        ws["I2"] = "FITNESS"
        ws["J2"] = "COVERAGE"
        ws["K2"] = "BALANCE"
        ws["L2"] = "FAIRNESS"
        ws["M2"] = "AVG"

        ws["I3"] = stat[0]
        ws["J3"] = stat[1]
        ws["K3"] = stat[2]
        ws["L3"] = stat[3]
        ws["M3"] = stat[4]

        wb.save(res_file)

        with pd.ExcelWriter(res_file, mode='a') as writer:
            stat[5].to_excel(writer, sheet_name='Blad2')


def run_different_dbs( db_dir):
    for db_file in [x for x in os.listdir(db_dir) if x[0] != "."]:

        db_name = db_dir[17:] + db_file 
        print(db_name)
        rel_dir_name = db_file.split(".")[0]
        dir_name = "/Users/dennismaes/Library/Mobile Documents/com~apple~CloudDocs/School/2021-2022/Thesis/applicatie/experiments/results_diff_periods/" + rel_dir_name + "/"
        run_all_custom_db(dir_name, db_name)


def run_different_dbs( db_dir):
    for db_file in [x for x in os.listdir(db_dir) if x[0] != "."]:

        db_name = db_dir[17:] + db_file 
        print(db_name)
        rel_dir_name = db_file.split(".")[0]
        dir_name = "/Users/dennismaes/Library/Mobile Documents/com~apple~CloudDocs/School/2021-2022/Thesis/applicatie/experiments/results_diff_periods/" + rel_dir_name + "/"
        run_all_custom_db(dir_name, db_name)


def run_different_dbs_no_subpopulations( db_dir, new_random=False):
    for db_file in [x for x in os.listdir(db_dir) if x[0] != "."]:

        db_name = db_dir[17:] + db_file 
        print(db_name)
        rel_dir_name = db_file.split(".")[0]
        dir_name = "/Users/dennismaes/Library/Mobile Documents/com~apple~CloudDocs/School/2021-2022/Thesis/applicatie/experiments/results_no_subpopulations/" + rel_dir_name + "/"
        print(dir_name)
        run_no_subpopulations(dir_name, db_name, new_random)

def run_different_dbs_no_subpopulations_new_random( db_dir):
    for db_file in [x for x in os.listdir(db_dir) if x[0] != "."]:

        db_name = db_dir[17:] + db_file 
        print(db_name)
        rel_dir_name = db_file.split(".")[0]
        dir_name = "/Users/dennismaes/Library/Mobile Documents/com~apple~CloudDocs/School/2021-2022/Thesis/applicatie/experiments/results_no_subpopulations_new_random/" + rel_dir_name + "/"
        print(dir_name)
        run_no_subpopulations(dir_name, db_name, new_random=True)


def run_different_dbs_higher_population_size( db_dir):
    for db_file in [x for x in os.listdir(db_dir) if x[0] != "."]:

        db_name = db_dir[17:] + db_file 
        print(db_name)
        rel_dir_name = db_file.split(".")[0]
        dir_name = "/Users/dennismaes/Library/Mobile Documents/com~apple~CloudDocs/School/2021-2022/Thesis/applicatie/experiments/results_higher_population_size/" + rel_dir_name + "/"
        print(dir_name)
        run_all_custom_db(dir_name, db_name)


def run_different_dbs_lower_survival_rate(db_dir):
    for db_file in [x for x in os.listdir(db_dir) if x[0] != "."]:

        db_name = db_dir[17:] + db_file 
        print(db_name)
        rel_dir_name = db_file.split(".")[0]
        dir_name = "/Users/dennismaes/Library/Mobile Documents/com~apple~CloudDocs/School/2021-2022/Thesis/applicatie/experiments/results_lower_survival_rate/" + rel_dir_name + "/"
        print(dir_name)
        run_all_custom_db(dir_name, db_name)


def run_different_dbs_test_case(db_dir):
    for db_file in [x for x in os.listdir(db_dir) if x[0] != "."]:

        db_name = db_dir[17:] + db_file 
        print(db_name)
        rel_dir_name = db_file.split(".")[0]
        dir_name = "/Users/dennismaes/Library/Mobile Documents/com~apple~CloudDocs/School/2021-2022/Thesis/applicatie/experiments/results_test_case/" + rel_dir_name + "/"
        print(dir_name)
        run_all_custom_db(dir_name, db_name)


def run_different_dbs_test_case2(db_dir):
    for db_file in [x for x in os.listdir(db_dir) if x[0] != "."]:

        db_name = db_dir[17:] + db_file 
        print(db_name)
        rel_dir_name = db_file.split(".")[0]
        dir_name = "/Users/dennismaes/Library/Mobile Documents/com~apple~CloudDocs/School/2021-2022/Thesis/applicatie/experiments/results_test_case2/" + rel_dir_name + "/"
        print(dir_name)
        run_all_custom_db(dir_name, db_name)


def run_no_subpopulations(dir_name, db_name, new_random=False):
    arr = os.listdir(dir_name)
    arr = [a for a in arr if a[0:3] == 'one']
    counter = 0
    for res_file in arr:
        counter += 1
        res_file = dir_name + res_file
        print("CURRENT FILE: ", res_file)
        print("STATUS: ", counter, "/", len(arr))

        wb = px.load_workbook(res_file)
        sh_n = "Blad1"
        print("Current sheetname: ", sh_n)
        ws = wb[sh_n]


        run_string = "java -cp " + jar_file + " Main"

        

        input_params = parse_params(wb)
        for x in input_params:
            run_string += " \"{}\"".format(x)

        input_data = parse_excel_input_no_subpopulations(ws)
        for x in input_data:
            run_string += " \"{}\"".format(x)

        run_string += " "
        run_string += " \"" + db_name + "\"" 


        if new_random:
            run_string += " newRandom"
        print(run_string)
        run_string += " > tmp.txt"

        os.system(run_string)

        

        schedule_entries = []
        for row in cur.execute('''SELECT assistant_id, day_nb, shift_type FROM assignment'''):
            schedule_entries.append(row)

        assistant_entries = dict()
        for row in cur.execute('''SELECT id, name FROM assistant'''):
            assistant_entries[row[0]] = row[1].strip()

        results_list = list()
        results7 = dict()
        results12 = dict()
        results14 = dict()
        results21 = dict()
        for as_id in list(assistant_entries.keys()):
            current_assignments = [x for x in schedule_entries if x[0] == as_id]
            current_name = assistant_entries[as_id]
            shifts = [x[2] for x in current_assignments]
            i = 0
            free_cnt = 0
            while i < len(shifts):
                if shifts[i] == 'FREE':
                    free_cnt += 1
                else:
                    if free_cnt == 2:
                        shifts[i - 2] = 'JAEV'
                        shifts[i - 1] = 'JAEV'
                    free_cnt = 0
                i += 1

            results7[current_name] = (get_nb_times_rest_less_than(shifts, 7), get_total_wl(shifts))
            results12[current_name] = (get_nb_times_rest_less_than(shifts, 12), get_total_wl(shifts))
            results14[current_name] = (get_nb_times_rest_less_than(shifts, 14), get_total_wl(shifts))
            results21[current_name] = (get_nb_times_rest_less_than(shifts, 21), get_total_wl(shifts))
        results_list.append(results7)
        results_list.append(results12)
        results_list.append(results14)
        results_list.append(results21)

        write_to_excel_ws(results_list, wb, ws, res_file)
        stat = read_results2("tmp.txt")

        ws["I2"] = "FITNESS"
        ws["J2"] = "COVERAGE"
        ws["K2"] = "BALANCE"
        ws["L2"] = "FAIRNESS"
        ws["M2"] = "AVG"

        ws["I3"] = stat[0]
        ws["J3"] = stat[1]
        ws["K3"] = stat[2]
        ws["L3"] = stat[3]
        ws["M3"] = stat[4]

        wb.save(res_file)

        with pd.ExcelWriter(res_file, mode='a') as writer:
            stat[5].to_excel(writer, sheet_name='Blad2')

if __name__ == "__main__":
    args = sys.argv
    if len(args) > 1:
        if args[1] == "different_months":
            run_different_dbs("/Users/dennismaes/Library/Mobile Documents/com~apple~CloudDocs/School/2021-2022/Thesis/applicatie/experiments/dbs/")
        
        elif args[1] == "no_subpopulations":
            if len(args) > 2:
                if args[2] == "random":
                    run_different_dbs_no_subpopulations_new_random("/Users/dennismaes/Library/Mobile Documents/com~apple~CloudDocs/School/2021-2022/Thesis/applicatie/experiments/dbs_no_subpopulations_new_random/")   

            else:     
                run_different_dbs_no_subpopulations("/Users/dennismaes/Library/Mobile Documents/com~apple~CloudDocs/School/2021-2022/Thesis/applicatie/experiments/dbs_no_subpopulations/")
        elif args[1] == "higher_population_size":
            run_different_dbs_higher_population_size("/Users/dennismaes/Library/Mobile Documents/com~apple~CloudDocs/School/2021-2022/Thesis/applicatie/experiments/dbs_higher_population_size/")
        
        elif args[1] == "lower_survival_rate":
            run_different_dbs_lower_survival_rate("/Users/dennismaes/Library/Mobile Documents/com~apple~CloudDocs/School/2021-2022/Thesis/applicatie/experiments/dbs_lower_survival_rate/")
        
        elif args[1] == "test_case":
            run_different_dbs_test_case("/Users/dennismaes/Library/Mobile Documents/com~apple~CloudDocs/School/2021-2022/Thesis/applicatie/experiments/dbs_test_case/")
        
        elif args[1] == "test_case2":
            run_different_dbs_test_case2("/Users/dennismaes/Library/Mobile Documents/com~apple~CloudDocs/School/2021-2022/Thesis/applicatie/experiments/dbs_test_case2/")
    else:
        run_all("/Users/dennismaes/Library/Mobile Documents/com~apple~CloudDocs/School/2021-2022/Thesis/applicatie/experiments/results6/")
            
        
    #run_all("/Users/dennismaes/Library/Mobile Documents/com~apple~CloudDocs/School/2021-2022/Thesis/applicatie/experiments/results2/")
    """cur.execute('''DELETE FROM assistant WHERE id = 73''')
    conn.commit()
    conn.close()"""
    